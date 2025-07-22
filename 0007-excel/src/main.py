from pathlib import Path
from pprint import pprint

import pandas as pd
from openpyxl import load_workbook

DATAPATH = Path(__file__).parent.parent.resolve() / "data" / "temp_data.xlsx"


def try_openpyxl():
    workbook = load_workbook(DATAPATH)
    active_sheet = workbook.active
    print(f"{'Row': <7}: {active_sheet.max_row}")
    print(f"{'Column': <7}: {active_sheet.max_column}")
    for index in range(1, active_sheet.max_column + 1):
        print(f"{active_sheet.cell(row=1, column=index).value}")
    for index in range(1, active_sheet.max_row + 1):
        print(f"{active_sheet.cell(row=index, column=1).value}")


def try_pandas():
    df = pd.read_excel(DATAPATH)
    data_dict = df.to_dict(orient="records")
    pprint(data_dict)


def main():
    try_pandas()


if __name__ == "__main__":
    main()
